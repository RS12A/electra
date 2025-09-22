"""
Export utilities for analytics data.

This module provides functions for exporting analytics data in various
formats (CSV, XLSX, PDF) with verification and security features.
"""
import csv
import hashlib
import io
import json
from datetime import datetime
from typing import Dict, Any, List, Union, ByteString
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth import get_user_model

# Import optional dependencies for XLSX and PDF
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .models import ExportVerification

User = get_user_model()


class AnalyticsExporter:
    """
    Main exporter class for analytics data with verification support.
    """
    
    def __init__(self, user: User, request_ip: str):
        """
        Initialize the exporter.
        
        Args:
            user: User requesting the export
            request_ip: IP address of the request
        """
        self.user = user
        self.request_ip = request_ip
        self.timestamp = timezone.now()
    
    def export_data(
        self,
        data: Dict[str, Any],
        export_type: str,
        filename: str,
        export_params: Dict[str, Any]
    ) -> tuple[bytes, ExportVerification]:
        """
        Export data in the specified format with verification.
        
        Args:
            data: Analytics data to export
            export_type: Format (csv, xlsx, pdf)
            filename: Base filename for the export
            export_params: Parameters used for the export
            
        Returns:
            Tuple of (exported_data_bytes, verification_record)
        """
        # Generate appropriate filename with timestamp
        timestamp_str = self.timestamp.strftime('%Y%m%d_%H%M%S')
        full_filename = f"{filename}_{timestamp_str}.{export_type}"
        
        # Export based on type
        if export_type == 'csv':
            exported_data = self._export_csv(data, full_filename)
        elif export_type == 'xlsx':
            if not XLSX_AVAILABLE:
                raise ValueError("XLSX export not available. Please install openpyxl.")
            exported_data = self._export_xlsx(data, full_filename)
        elif export_type == 'pdf':
            if not PDF_AVAILABLE:
                raise ValueError("PDF export not available. Please install reportlab.")
            exported_data = self._export_pdf(data, full_filename)
        else:
            raise ValueError(f"Unsupported export type: {export_type}")
        
        # Create verification record
        verification = ExportVerification.create_for_export(
            export_type=export_type,
            content=exported_data,
            filename=full_filename,
            export_params=export_params,
            requested_by=self.user,
            request_ip=self.request_ip
        )
        
        return exported_data, verification
    
    def _export_csv(self, data: Dict[str, Any], filename: str) -> bytes:
        """Export data as CSV format."""
        output = io.StringIO()
        
        # Add header with metadata
        writer = csv.writer(output)
        writer.writerow(['# Electra Analytics Export'])
        writer.writerow(['# Generated:', self.timestamp.isoformat()])
        writer.writerow(['# Exported by:', self.user.full_name])
        writer.writerow(['# Export type:', 'CSV'])
        writer.writerow([])  # Empty row
        
        # Determine data structure and export accordingly
        if 'per_election' in data:  # Turnout data
            self._export_turnout_csv(writer, data)
        elif 'by_user_type' in data:  # Participation data
            self._export_participation_csv(writer, data)
        elif 'data_points' in data:  # Time series data
            self._export_time_series_csv(writer, data)
        elif 'election' in data:  # Election summary
            self._export_election_summary_csv(writer, data)
        else:
            # Generic data export
            self._export_generic_csv(writer, data)
        
        # Add verification footer
        writer.writerow([])
        content_hash = hashlib.sha256(output.getvalue().encode()).hexdigest()
        writer.writerow(['# Content Hash:', content_hash])
        
        return output.getvalue().encode('utf-8')
    
    def _export_turnout_csv(self, writer, data: Dict[str, Any]) -> None:
        """Export turnout data to CSV."""
        writer.writerow(['TURNOUT METRICS'])
        writer.writerow(['Overall Turnout:', f"{data.get('overall_turnout', 0):.2f}%"])
        writer.writerow([])
        
        # Summary section
        summary = data.get('summary', {})
        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Elections:', summary.get('total_elections', 0)])
        writer.writerow(['Active Elections:', summary.get('active_elections', 0)])
        writer.writerow(['Completed Elections:', summary.get('completed_elections', 0)])
        writer.writerow(['Total Eligible Voters:', summary.get('total_eligible_voters', 0)])
        writer.writerow(['Total Votes Cast:', summary.get('total_votes_cast', 0)])
        writer.writerow([])
        
        # Per-election data
        writer.writerow(['PER-ELECTION DETAILS'])
        writer.writerow([
            'Election ID', 'Title', 'Status', 'Eligible Voters',
            'Votes Cast', 'Turnout %', 'Category', 'Start Time', 'End Time'
        ])
        
        for election in data.get('per_election', []):
            writer.writerow([
                election.get('election_id', ''),
                election.get('election_title', ''),
                election.get('status', ''),
                election.get('eligible_voters', 0),
                election.get('votes_cast', 0),
                f"{election.get('turnout_percentage', 0):.2f}",
                election.get('category', ''),
                election.get('start_time', ''),
                election.get('end_time', ''),
            ])
    
    def _export_participation_csv(self, writer, data: Dict[str, Any]) -> None:
        """Export participation data to CSV."""
        writer.writerow(['PARTICIPATION ANALYTICS'])
        writer.writerow([])
        
        # Summary
        summary = data.get('summary', {})
        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Eligible Users:', summary.get('total_eligible_users', 0)])
        writer.writerow(['Total Participants:', summary.get('total_participants', 0)])
        writer.writerow(['Overall Participation Rate:', f"{summary.get('overall_participation_rate', 0):.2f}%"])
        writer.writerow([])
        
        # By user type
        writer.writerow(['BY USER TYPE'])
        writer.writerow(['User Type', 'Eligible Users', 'Participants', 'Participation Rate %', 'Category'])
        
        for user_type, type_data in data.get('by_user_type', {}).items():
            writer.writerow([
                user_type,
                type_data.get('eligible_users', 0),
                type_data.get('participants', 0),
                f"{type_data.get('participation_rate', 0):.2f}",
                type_data.get('category', ''),
            ])
        
        writer.writerow([])
        
        # By category
        writer.writerow(['BY CATEGORY'])
        writer.writerow(['Category', 'Count'])
        for category, count in data.get('by_category', {}).items():
            writer.writerow([category, count])
    
    def _export_time_series_csv(self, writer, data: Dict[str, Any]) -> None:
        """Export time series data to CSV."""
        writer.writerow(['TIME SERIES ANALYTICS'])
        writer.writerow(['Period Type:', data.get('period_type', '')])
        writer.writerow(['Start Date:', data.get('start_date', '')])
        writer.writerow(['End Date:', data.get('end_date', '')])
        writer.writerow([])
        
        # Summary
        summary = data.get('summary', {})
        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Votes:', summary.get('total_votes', 0)])
        writer.writerow(['Average Daily Votes:', summary.get('average_daily_votes', 0)])
        
        peak_day = summary.get('peak_voting_day')
        if peak_day:
            writer.writerow(['Peak Voting Day:', peak_day.get('period', '')])
            writer.writerow(['Peak Day Votes:', peak_day.get('vote_count', 0)])
        
        writer.writerow([])
        
        # Data points
        writer.writerow(['TIME SERIES DATA'])
        writer.writerow(['Period', 'Vote Count', 'Period Start', 'Period End'])
        
        for point in data.get('data_points', []):
            writer.writerow([
                point.get('period', ''),
                point.get('vote_count', 0),
                point.get('period_start', ''),
                point.get('period_end', ''),
            ])
    
    def _export_election_summary_csv(self, writer, data: Dict[str, Any]) -> None:
        """Export election summary to CSV."""
        election = data.get('election', {})
        writer.writerow(['ELECTION SUMMARY'])
        writer.writerow(['Election ID:', election.get('id', '')])
        writer.writerow(['Title:', election.get('title', '')])
        writer.writerow(['Status:', election.get('status', '')])
        writer.writerow(['Start Time:', election.get('start_time', '')])
        writer.writerow(['End Time:', election.get('end_time', '')])
        writer.writerow([])
        
        # Include turnout data
        turnout = data.get('turnout', {})
        if turnout:
            writer.writerow(['TURNOUT'])
            writer.writerow(['Eligible Voters:', turnout.get('eligible_voters', 0)])
            writer.writerow(['Votes Cast:', turnout.get('votes_cast', 0)])
            writer.writerow(['Turnout Percentage:', f"{turnout.get('turnout_percentage', 0):.2f}%"])
            writer.writerow(['Category:', turnout.get('category', '')])
            writer.writerow([])
        
        # Include participation summary
        participation = data.get('participation', {})
        if participation:
            part_summary = participation.get('summary', {})
            writer.writerow(['PARTICIPATION'])
            writer.writerow(['Total Eligible Users:', part_summary.get('total_eligible_users', 0)])
            writer.writerow(['Total Participants:', part_summary.get('total_participants', 0)])
            writer.writerow(['Participation Rate:', f"{part_summary.get('overall_participation_rate', 0):.2f}%"])
    
    def _export_generic_csv(self, writer, data: Dict[str, Any]) -> None:
        """Export generic data structure to CSV."""
        def flatten_dict(d, parent_key='', sep='_'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    for i, item in enumerate(v):
                        if isinstance(item, dict):
                            items.extend(flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
                        else:
                            items.append((f"{new_key}_{i}", item))
                else:
                    items.append((new_key, v))
            return dict(items)
        
        flattened = flatten_dict(data)
        writer.writerow(['Key', 'Value'])
        for key, value in flattened.items():
            writer.writerow([key, str(value)])
    
    def _export_xlsx(self, data: Dict[str, Any], filename: str) -> bytes:
        """Export data as XLSX format with formatting."""
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # Remove default sheet and create new ones
        workbook.remove(workbook.active)
        
        # Determine data structure and create appropriate sheets
        if 'per_election' in data:  # Turnout data
            self._create_turnout_xlsx_sheets(workbook, data)
        elif 'by_user_type' in data:  # Participation data
            self._create_participation_xlsx_sheets(workbook, data)
        elif 'data_points' in data:  # Time series data
            self._create_time_series_xlsx_sheets(workbook, data)
        elif 'election' in data:  # Election summary
            self._create_election_summary_xlsx_sheets(workbook, data)
        else:
            # Generic data
            self._create_generic_xlsx_sheet(workbook, data)
        
        # Add metadata sheet
        self._add_metadata_sheet(workbook)
        
        workbook.save(output)
        return output.getvalue()
    
    def _create_turnout_xlsx_sheets(self, workbook, data: Dict[str, Any]) -> None:
        """Create XLSX sheets for turnout data."""
        # Summary sheet
        summary_sheet = workbook.create_sheet("Summary", 0)
        summary_sheet['A1'] = "Electra Analytics - Turnout Report"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        
        summary_data = data.get('summary', {})
        summary_sheet['A3'] = "Overall Statistics"
        summary_sheet['A3'].font = Font(bold=True)
        
        summary_sheet['A4'] = "Overall Turnout:"
        summary_sheet['B4'] = f"{data.get('overall_turnout', 0):.2f}%"
        summary_sheet['A5'] = "Total Elections:"
        summary_sheet['B5'] = summary_data.get('total_elections', 0)
        summary_sheet['A6'] = "Active Elections:"
        summary_sheet['B6'] = summary_data.get('active_elections', 0)
        summary_sheet['A7'] = "Completed Elections:"
        summary_sheet['B7'] = summary_data.get('completed_elections', 0)
        summary_sheet['A8'] = "Total Eligible Voters:"
        summary_sheet['B8'] = summary_data.get('total_eligible_voters', 0)
        summary_sheet['A9'] = "Total Votes Cast:"
        summary_sheet['B9'] = summary_data.get('total_votes_cast', 0)
        
        # Per-election details sheet
        details_sheet = workbook.create_sheet("Election Details")
        headers = [
            'Election ID', 'Title', 'Status', 'Eligible Voters',
            'Votes Cast', 'Turnout %', 'Category', 'Start Time', 'End Time'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = details_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, election in enumerate(data.get('per_election', []), 2):
            details_sheet.cell(row=row, column=1, value=election.get('election_id', ''))
            details_sheet.cell(row=row, column=2, value=election.get('election_title', ''))
            details_sheet.cell(row=row, column=3, value=election.get('status', ''))
            details_sheet.cell(row=row, column=4, value=election.get('eligible_voters', 0))
            details_sheet.cell(row=row, column=5, value=election.get('votes_cast', 0))
            details_sheet.cell(row=row, column=6, value=election.get('turnout_percentage', 0))
            details_sheet.cell(row=row, column=7, value=election.get('category', ''))
            details_sheet.cell(row=row, column=8, value=election.get('start_time', ''))
            details_sheet.cell(row=row, column=9, value=election.get('end_time', ''))
        
        # Auto-adjust column widths
        for sheet in [summary_sheet, details_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_participation_xlsx_sheets(self, workbook, data: Dict[str, Any]) -> None:
        """Create XLSX sheets for participation data."""
        # Summary sheet
        summary_sheet = workbook.create_sheet("Summary", 0)
        summary_sheet['A1'] = "Electra Analytics - Participation Report"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        
        summary_data = data.get('summary', {})
        summary_sheet['A3'] = "Overall Statistics"
        summary_sheet['A3'].font = Font(bold=True)
        
        summary_sheet['A4'] = "Total Eligible Users:"
        summary_sheet['B4'] = summary_data.get('total_eligible_users', 0)
        summary_sheet['A5'] = "Total Participants:"
        summary_sheet['B5'] = summary_data.get('total_participants', 0)
        summary_sheet['A6'] = "Overall Participation Rate:"
        summary_sheet['B6'] = f"{summary_data.get('overall_participation_rate', 0):.2f}%"
        
        # By user type sheet
        user_type_sheet = workbook.create_sheet("By User Type")
        headers = ['User Type', 'Eligible Users', 'Participants', 'Participation Rate %', 'Category']
        
        for col, header in enumerate(headers, 1):
            cell = user_type_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 2
        for user_type, type_data in data.get('by_user_type', {}).items():
            user_type_sheet.cell(row=row, column=1, value=user_type)
            user_type_sheet.cell(row=row, column=2, value=type_data.get('eligible_users', 0))
            user_type_sheet.cell(row=row, column=3, value=type_data.get('participants', 0))
            user_type_sheet.cell(row=row, column=4, value=type_data.get('participation_rate', 0))
            user_type_sheet.cell(row=row, column=5, value=type_data.get('category', ''))
            row += 1
        
        # Auto-adjust column widths
        for sheet in [summary_sheet, user_type_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_time_series_xlsx_sheets(self, workbook, data: Dict[str, Any]) -> None:
        """Create XLSX sheets for time series data."""
        # Summary sheet
        summary_sheet = workbook.create_sheet("Summary", 0)
        summary_sheet['A1'] = "Electra Analytics - Time Series Report"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        
        summary_sheet['A3'] = "Report Parameters"
        summary_sheet['A3'].font = Font(bold=True)
        summary_sheet['A4'] = "Period Type:"
        summary_sheet['B4'] = data.get('period_type', '')
        summary_sheet['A5'] = "Start Date:"
        summary_sheet['B5'] = data.get('start_date', '')
        summary_sheet['A6'] = "End Date:"
        summary_sheet['B6'] = data.get('end_date', '')
        
        summary_data = data.get('summary', {})
        summary_sheet['A8'] = "Summary Statistics"
        summary_sheet['A8'].font = Font(bold=True)
        summary_sheet['A9'] = "Total Votes:"
        summary_sheet['B9'] = summary_data.get('total_votes', 0)
        summary_sheet['A10'] = "Average Daily Votes:"
        summary_sheet['B10'] = summary_data.get('average_daily_votes', 0)
        
        peak_day = summary_data.get('peak_voting_day')
        if peak_day:
            summary_sheet['A11'] = "Peak Voting Day:"
            summary_sheet['B11'] = peak_day.get('period', '')
            summary_sheet['A12'] = "Peak Day Votes:"
            summary_sheet['B12'] = peak_day.get('vote_count', 0)
        
        # Data sheet
        data_sheet = workbook.create_sheet("Time Series Data")
        headers = ['Period', 'Vote Count', 'Period Start', 'Period End']
        
        for col, header in enumerate(headers, 1):
            cell = data_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, point in enumerate(data.get('data_points', []), 2):
            data_sheet.cell(row=row, column=1, value=point.get('period', ''))
            data_sheet.cell(row=row, column=2, value=point.get('vote_count', 0))
            data_sheet.cell(row=row, column=3, value=point.get('period_start', ''))
            data_sheet.cell(row=row, column=4, value=point.get('period_end', ''))
        
        # Create chart if data exists
        if data.get('data_points'):
            chart = BarChart()
            chart.title = "Votes Over Time"
            chart.x_axis.title = "Time Period"
            chart.y_axis.title = "Vote Count"
            
            data_range = Reference(data_sheet, min_col=2, min_row=1, max_row=len(data['data_points'])+1)
            categories = Reference(data_sheet, min_col=1, min_row=2, max_row=len(data['data_points'])+1)
            
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(categories)
            data_sheet.add_chart(chart, "F5")
        
        # Auto-adjust column widths
        for sheet in [summary_sheet, data_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_election_summary_xlsx_sheets(self, workbook, data: Dict[str, Any]) -> None:
        """Create XLSX sheets for election summary data."""
        # Main summary sheet
        summary_sheet = workbook.create_sheet("Election Summary", 0)
        summary_sheet['A1'] = "Electra Analytics - Election Summary"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        
        election = data.get('election', {})
        summary_sheet['A3'] = "Election Information"
        summary_sheet['A3'].font = Font(bold=True)
        summary_sheet['A4'] = "Election ID:"
        summary_sheet['B4'] = election.get('id', '')
        summary_sheet['A5'] = "Title:"
        summary_sheet['B5'] = election.get('title', '')
        summary_sheet['A6'] = "Status:"
        summary_sheet['B6'] = election.get('status', '')
        summary_sheet['A7'] = "Start Time:"
        summary_sheet['B7'] = election.get('start_time', '')
        summary_sheet['A8'] = "End Time:"
        summary_sheet['B8'] = election.get('end_time', '')
        
        # Turnout information
        turnout = data.get('turnout', {})
        if turnout:
            summary_sheet['A10'] = "Turnout Information"
            summary_sheet['A10'].font = Font(bold=True)
            summary_sheet['A11'] = "Eligible Voters:"
            summary_sheet['B11'] = turnout.get('eligible_voters', 0)
            summary_sheet['A12'] = "Votes Cast:"
            summary_sheet['B12'] = turnout.get('votes_cast', 0)
            summary_sheet['A13'] = "Turnout Percentage:"
            summary_sheet['B13'] = f"{turnout.get('turnout_percentage', 0):.2f}%"
            summary_sheet['A14'] = "Category:"
            summary_sheet['B14'] = turnout.get('category', '')
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_generic_xlsx_sheet(self, workbook, data: Dict[str, Any]) -> None:
        """Create a generic XLSX sheet for unstructured data."""
        sheet = workbook.create_sheet("Data", 0)
        sheet['A1'] = "Electra Analytics Export"
        sheet['A1'].font = Font(size=16, bold=True)
        
        # Convert data to JSON string for display
        sheet['A3'] = "Raw Data:"
        sheet['A3'].font = Font(bold=True)
        sheet['A4'] = json.dumps(data, indent=2, default=str)
    
    def _add_metadata_sheet(self, workbook) -> None:
        """Add metadata sheet to workbook."""
        metadata_sheet = workbook.create_sheet("Metadata")
        metadata_sheet['A1'] = "Export Metadata"
        metadata_sheet['A1'].font = Font(size=14, bold=True)
        
        metadata_sheet['A3'] = "Generated:"
        metadata_sheet['B3'] = self.timestamp.isoformat()
        metadata_sheet['A4'] = "Exported by:"
        metadata_sheet['B4'] = self.user.full_name
        metadata_sheet['A5'] = "User Role:"
        metadata_sheet['B5'] = self.user.role
        metadata_sheet['A6'] = "Export Type:"
        metadata_sheet['B6'] = "Excel (XLSX)"
        
        # Auto-adjust column widths
        for column in metadata_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            metadata_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _export_pdf(self, data: Dict[str, Any], filename: str) -> bytes:
        """Export data as PDF format."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Electra Analytics Report", title_style))
        story.append(Spacer(1, 12))
        
        # Metadata
        story.append(Paragraph(f"<b>Generated:</b> {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Paragraph(f"<b>Exported by:</b> {self.user.full_name}", styles['Normal']))
        story.append(Paragraph(f"<b>User Role:</b> {self.user.role}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Determine data structure and add appropriate content
        if 'per_election' in data:  # Turnout data
            self._add_turnout_pdf_content(story, data, styles)
        elif 'by_user_type' in data:  # Participation data
            self._add_participation_pdf_content(story, data, styles)
        elif 'data_points' in data:  # Time series data
            self._add_time_series_pdf_content(story, data, styles)
        elif 'election' in data:  # Election summary
            self._add_election_summary_pdf_content(story, data, styles)
        else:
            # Generic data
            self._add_generic_pdf_content(story, data, styles)
        
        # Add verification footer
        story.append(Spacer(1, 30))
        content_hash = hashlib.sha256(json.dumps(data, sort_keys=True).encode()).hexdigest()
        story.append(Paragraph(f"<b>Content Verification Hash:</b> {content_hash[:32]}...", styles['Normal']))
        
        doc.build(story)
        return output.getvalue()
    
    def _add_turnout_pdf_content(self, story, data: Dict[str, Any], styles) -> None:
        """Add turnout data content to PDF."""
        story.append(Paragraph("Turnout Metrics Report", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Overall turnout
        story.append(Paragraph(f"<b>Overall Turnout:</b> {data.get('overall_turnout', 0):.2f}%", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Summary table
        summary = data.get('summary', {})
        summary_data = [
            ['Metric', 'Value'],
            ['Total Elections', str(summary.get('total_elections', 0))],
            ['Active Elections', str(summary.get('active_elections', 0))],
            ['Completed Elections', str(summary.get('completed_elections', 0))],
            ['Total Eligible Voters', str(summary.get('total_eligible_voters', 0))],
            ['Total Votes Cast', str(summary.get('total_votes_cast', 0))],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Per-election details (limit to first 10 for space)
        if data.get('per_election'):
            story.append(Paragraph("Election Details", styles['Heading3']))
            story.append(Spacer(1, 12))
            
            election_data = [['Title', 'Status', 'Eligible', 'Votes', 'Turnout %', 'Category']]
            
            for election in data['per_election'][:10]:  # Limit for PDF space
                election_data.append([
                    election.get('election_title', '')[:30],  # Truncate for space
                    election.get('status', ''),
                    str(election.get('eligible_voters', 0)),
                    str(election.get('votes_cast', 0)),
                    f"{election.get('turnout_percentage', 0):.1f}%",
                    election.get('category', '')
                ])
            
            election_table = Table(election_data)
            election_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(election_table)
            
            if len(data['per_election']) > 10:
                story.append(Spacer(1, 12))
                story.append(Paragraph(f"<i>Note: Showing first 10 of {len(data['per_election'])} elections. Full data available in CSV/XLSX export.</i>", styles['Normal']))
    
    def _add_participation_pdf_content(self, story, data: Dict[str, Any], styles) -> None:
        """Add participation data content to PDF."""
        story.append(Paragraph("Participation Analytics Report", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Summary
        summary = data.get('summary', {})
        story.append(Paragraph(f"<b>Total Eligible Users:</b> {summary.get('total_eligible_users', 0)}", styles['Normal']))
        story.append(Paragraph(f"<b>Total Participants:</b> {summary.get('total_participants', 0)}", styles['Normal']))
        story.append(Paragraph(f"<b>Overall Participation Rate:</b> {summary.get('overall_participation_rate', 0):.2f}%", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # By user type
        if data.get('by_user_type'):
            story.append(Paragraph("Participation by User Type", styles['Heading3']))
            story.append(Spacer(1, 12))
            
            user_type_data = [['User Type', 'Eligible', 'Participants', 'Rate %', 'Category']]
            
            for user_type, type_data in data['by_user_type'].items():
                user_type_data.append([
                    user_type,
                    str(type_data.get('eligible_users', 0)),
                    str(type_data.get('participants', 0)),
                    f"{type_data.get('participation_rate', 0):.1f}%",
                    type_data.get('category', '')
                ])
            
            user_type_table = Table(user_type_data)
            user_type_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(user_type_table)
    
    def _add_time_series_pdf_content(self, story, data: Dict[str, Any], styles) -> None:
        """Add time series data content to PDF."""
        story.append(Paragraph("Time Series Analytics Report", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Parameters
        story.append(Paragraph(f"<b>Period Type:</b> {data.get('period_type', '')}", styles['Normal']))
        story.append(Paragraph(f"<b>Date Range:</b> {data.get('start_date', '')} to {data.get('end_date', '')}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Summary
        summary = data.get('summary', {})
        story.append(Paragraph(f"<b>Total Votes:</b> {summary.get('total_votes', 0)}", styles['Normal']))
        story.append(Paragraph(f"<b>Average Daily Votes:</b> {summary.get('average_daily_votes', 0)}", styles['Normal']))
        
        peak_day = summary.get('peak_voting_day')
        if peak_day:
            story.append(Paragraph(f"<b>Peak Voting Day:</b> {peak_day.get('period', '')} ({peak_day.get('vote_count', 0)} votes)", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Data points (limit for PDF space)
        if data.get('data_points'):
            story.append(Paragraph("Time Series Data", styles['Heading3']))
            story.append(Spacer(1, 12))
            
            data_points = data['data_points'][:20]  # Limit for PDF space
            time_series_data = [['Period', 'Vote Count']]
            
            for point in data_points:
                time_series_data.append([
                    point.get('period', ''),
                    str(point.get('vote_count', 0))
                ])
            
            time_series_table = Table(time_series_data)
            time_series_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(time_series_table)
            
            if len(data['data_points']) > 20:
                story.append(Spacer(1, 12))
                story.append(Paragraph(f"<i>Note: Showing first 20 of {len(data['data_points'])} data points. Full data available in CSV/XLSX export.</i>", styles['Normal']))
    
    def _add_election_summary_pdf_content(self, story, data: Dict[str, Any], styles) -> None:
        """Add election summary content to PDF."""
        story.append(Paragraph("Election Summary Report", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Election information
        election = data.get('election', {})
        story.append(Paragraph(f"<b>Election:</b> {election.get('title', '')}", styles['Normal']))
        story.append(Paragraph(f"<b>Status:</b> {election.get('status', '')}", styles['Normal']))
        story.append(Paragraph(f"<b>Period:</b> {election.get('start_time', '')} to {election.get('end_time', '')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Turnout summary
        turnout = data.get('turnout', {})
        if turnout:
            story.append(Paragraph("Turnout Summary", styles['Heading3']))
            story.append(Spacer(1, 12))
            
            turnout_data = [
                ['Metric', 'Value'],
                ['Eligible Voters', str(turnout.get('eligible_voters', 0))],
                ['Votes Cast', str(turnout.get('votes_cast', 0))],
                ['Turnout Percentage', f"{turnout.get('turnout_percentage', 0):.2f}%"],
                ['Category', turnout.get('category', '')],
            ]
            
            turnout_table = Table(turnout_data)
            turnout_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(turnout_table)
    
    def _add_generic_pdf_content(self, story, data: Dict[str, Any], styles) -> None:
        """Add generic data content to PDF."""
        story.append(Paragraph("Analytics Data Export", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Convert data to formatted text
        data_text = json.dumps(data, indent=2, default=str)
        
        # Split into chunks to avoid page overflow
        max_chars = 2000
        if len(data_text) > max_chars:
            data_text = data_text[:max_chars] + "\n\n... (truncated for PDF display)"
        
        story.append(Paragraph("<b>Raw Data:</b>", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Use a monospace style for JSON
        code_style = ParagraphStyle(
            'Code',
            parent=styles['Normal'],
            fontName='Courier',
            fontSize=8,
            leftIndent=20
        )
        
        for line in data_text.split('\n'):
            if line.strip():
                story.append(Paragraph(line, code_style))


def create_download_response(
    content: bytes,
    filename: str,
    content_type: str
) -> HttpResponse:
    """
    Create HTTP response for file download.
    
    Args:
        content: File content as bytes
        filename: Filename for download
        content_type: MIME content type
        
    Returns:
        HttpResponse configured for file download
    """
    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(content)
    
    # Add security headers
    response['X-Content-Type-Options'] = 'nosniff'
    response['X-Frame-Options'] = 'DENY'
    
    return response